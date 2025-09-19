from django.contrib import admin
from django.http import HttpResponse
from .models import AuditLog
import openpyxl
from openpyxl.utils import get_column_letter

@admin.register(AuditLog)
class AuditLogAdmin(admin.ModelAdmin):
    list_display = ("timestamp", "actor", "action", "content_type", "object_id", "short_desc")
    list_filter = ("action", "timestamp", "content_type")
    search_fields = ("description", "object_id", "actor__username")
    actions = ["export_to_excel"]

    def short_desc(self, obj):
        return (obj.description or "")[:80]

    @admin.action(description="Експорт у Excel")
    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit"

        headers = ["Дата/час", "Користувач", "Дія", "Модель", "ID об'єкта", "Опис", "Зміни (JSON)"]
        ws.append(headers)

        for log in queryset.order_by("timestamp"):
            ws.append([
                log.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                getattr(log.actor, "username", "") if log.actor else "",
                log.get_action_display(),
                f"{log.content_type.app_label}.{log.content_type.model}",
                str(log.object_id),
                log.description or "",
                "" if not log.changes else str(log.changes),
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="audit_export.xlsx"'
        wb.save(resp)
        return resp
