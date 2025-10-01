from __future__ import annotations
from django.contrib import admin
from django import forms
from django.http import HttpResponse
import io
import datetime

# Для експорту Excel
import openpyxl
from openpyxl.utils import get_column_letter

from .models import (
    QAEntry,
    UnansweredQuestion,
    Category,
    QuestionLog,
    AllowedTelegramUser,
    QAVariant,
)
from audittrail.admin_mixins import AuditedModelAdmin
from audittrail.models import AuditAction


# --------- QAEntry
@admin.register(QAEntry)
class QAEntryAdmin(AuditedModelAdmin):
    list_display = ("question", "category")
    list_filter = ("category",)
    search_fields = ("question", "synonyms", "answer")
    ordering = ("question",)
    list_per_page = 25

    # виключаємо embedding з audit (занадто великий масив)
    audit_exclude_fields = ("id", "embedding")

    def delete_queryset(self, request, queryset):
        # Логуємо кожен QAEntry перед видаленням (bulk)
        for obj in queryset:
            try:
                self.log_audit(
                    request,
                    obj,
                    AuditAction.DELETE,
                    description=f"Видалено {obj._meta.verbose_name} «{obj}» (bulk)"
                )
            except Exception:
                pass
        # Видаляємо варіанти пов'язані з цими QAEntry
        QAVariant.objects.filter(entry__in=queryset).delete()
        super().delete_queryset(request, queryset)

    def delete_model(self, request, obj):
        # Логуємо і видаляємо варіанти
        try:
            self.log_audit(
                request,
                obj,
                AuditAction.DELETE,
                description=f"Видалено {obj._meta.verbose_name} «{obj}»"
            )
        except Exception:
            pass
        QAVariant.objects.filter(entry=obj).delete()
        super().delete_model(request, obj)


# --------- Category
@admin.register(Category)
class CategoryAdmin(AuditedModelAdmin):
    search_fields = ('name',)


# --------- QuestionLog (експорт в Excel)
@admin.register(QuestionLog)
class QuestionLogAdmin(admin.ModelAdmin):
    list_display = ('question', 'answer_found', 'similarity', 'timestamp', 'asked_by')
    list_filter = ('answer_found', 'timestamp', 'asked_by')
    search_fields = (
        'question',
        'asked_by__full_name',
        'asked_by__user_id',
    )
    actions = ("export_to_excel",)

    def export_to_excel(self, request, queryset):
        """
        Експортує вибрані записи QuestionLog до Excel.
        Колонки: Timestamp, Question, Answer found, Similarity, Asked by (ПІБ), Asked by ID
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QuestionLog"

        headers = ["Timestamp", "Question", "Answer found", "Similarity", "Asked by (fullname)", "Asked by (user_id)"]
        ws.append(headers)

        for q in queryset.order_by("timestamp"):
            ts = q.timestamp.astimezone().strftime("%Y-%m-%d %H:%M:%S") if q.timestamp else ""
            asked_by_name = q.asked_by.full_name if getattr(q, "asked_by", None) else ""
            asked_by_id = q.asked_by.user_id if getattr(q, "asked_by", None) else ""
            row = [ts, q.question, "Yes" if q.answer_found else "No", q.similarity if q.similarity is not None else "", asked_by_name, asked_by_id]
            ws.append(row)

        # автопідігнати ширину колонок
        for i, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value or "")) for cell in column_cells) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(length, 60)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"question_logs_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    export_to_excel.short_description = "Експорт вибраних записів у Excel (.xlsx)"


# --------- UnansweredQuestion з додатковими полями (тепер успадковано від AuditedModelAdmin)
class UnansweredQuestionAdminForm(forms.ModelForm):
    synonyms = forms.CharField(
        label="Синоніми",
        required=False,
        help_text="Розділяй синоніми крапкою з комою ';'"
    )
    category = forms.ModelChoiceField(
        label="Категорія",
        queryset=Category.objects.all(),
        required=False
    )

    class Meta:
        model = UnansweredQuestion
        fields = ['question', 'proposed_answer', 'synonyms', 'category']


@admin.register(UnansweredQuestion)
class UnansweredQuestionAdmin(AuditedModelAdmin):
    form = UnansweredQuestionAdminForm
    list_display = ('question', 'asked_at', 'proposed_answer')
    search_fields = ('question',)
    date_hierarchy = "asked_at"

    def save_model(self, request, obj, form, change):
        """
        Якщо адміністратор заповнив proposed_answer — конвертуємо у QAEntry.
        Логуємо створення QAEntry та видалення UnansweredQuestion.
        """
        if obj.question and obj.proposed_answer:
            qa = QAEntry.objects.create(
                question=obj.question,
                answer=obj.proposed_answer,
                synonyms=form.cleaned_data.get('synonyms'),
                category=form.cleaned_data.get('category'),
            )
            # лог створення QAEntry
            try:
                self.log_audit(
                    request,
                    qa,
                    AuditAction.CREATE,
                    description=f"Створено {qa._meta.verbose_name} «{qa}» з UnansweredQuestion"
                )
            except Exception:
                pass

            # лог видалення UnansweredQuestion
            try:
                self.log_audit(
                    request,
                    obj,
                    AuditAction.DELETE,
                    description=f"Видалено {obj._meta.verbose_name} «{obj}» після конвертації в QAEntry"
                )
            except Exception:
                pass

            # видаляємо оригінал після логування
            obj.delete()
        else:
            super().save_model(request, obj, form, change)


# --------- AllowedTelegramUser
@admin.register(AllowedTelegramUser)
class AllowedTelegramUserAdmin(AuditedModelAdmin):
    list_display = ("user_id", "full_name", "status", "created_at")
    list_filter = ("status", "created_at")
    search_fields = ("user_id", "full_name")
    date_hierarchy = "created_at"
    ordering = ("-created_at",)
    list_per_page = 25
    actions = ("make_active", "make_inactive")

    @admin.action(description="Позначити як Активний")
    def make_active(self, request, queryset):
        updated = queryset.update(status=AllowedTelegramUser.Status.ACTIVE)
        self.message_user(request, f"Оновлено: {updated}")

    @admin.action(description="Позначити як Деактивований")
    def make_inactive(self, request, queryset):
        updated = queryset.update(status=AllowedTelegramUser.Status.INACTIVE)
        self.message_user(request, f"Оновлено: {updated}")
